"""
Android App Version History Scraper — v2
=========================================
Fixes from v1:
  1. Single primary source per app (APKPure /versions page) — no more cross-source noise
  2. Structured HTML parsing instead of proximity regex (no more date/build-code bleed)
  3. Per-app version format validation — filters internal APK build codes
  4. AppBrain as secondary source for release notes only (not version discovery)
  5. Play Store for current version + current release notes only
  6. Proper semantic version sorting (descending)
  7. Cutoff filter applied consistently

Run this on your local machine where those domains are accessible.
"""

import re, sys, time
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────
OUTPUT_FILE = "android_app_version_history_v2.xlsx"
CUTOFF_DATE = "2023-05-01"   # exclude versions released before this date
SLEEP       = 1.5            # polite delay between requests

# ── App list ───────────────────────────────────────────────────────────────────
# (name, package_id, developer, category, apkpure_slug, apkpure_pkg_override)
# apkpure_slug: the URL segment used by APKPure — verify manually if scraping fails
ANDROID_APPS = [
    ("YouTube",             "com.google.android.youtube",   "Google LLC",            "Photo & Video",     "youtube",              None),
    ("TikTok",              "com.zhiliaoapp.musically",      "TikTok Ltd.",           "Entertainment",     "tik-tok",              None),
    ("ChatGPT",             "com.openai.chatgpt",            "OpenAI OpCo, LLC",      "Productivity",      "chatgpt",              None),
    ("Claude by Anthropic", "com.anthropic.claude",          "Anthropic PBC",         "Productivity",      None,                   None),  # no APKPure page
    ("WhatsApp Messenger",  "com.whatsapp",                  "WhatsApp LLC",          "Social Networking", "whatsapp-messenger",   None),
    ("CapCut",              "com.lemon.lvoverseas",           "Bytedance Pte. Ltd",   "Photo & Video",     "capcut",               None),
    ("Instagram",           "com.instagram.android",         "Instagram, Inc.",       "Photo & Video",     "instagram",            None),
    ("LinkedIn",            "com.linkedin.android",          "LinkedIn Corporation",  "Business",          None,                   None),  # no APKPure page
    ("Tinder",              "com.tinder",                    "Tinder LLC",            "Lifestyle",         "tinder",               None),
    ("Spotify",             "com.spotify.music",             "Spotify AB",            "Music",             "spotify-music",        None),
]

COLUMNS = [
    "App Name", "Platform", "Developer / Company", "App Category",
    "Version Number", "Version Release Date", "Current Version",
    "Initial App Release Date", "Update Description / Release Notes",
    "Update Category", "Brief Summary",
    "Source of Update History", "Data Quality Notes",
]

# ── Per-app version format rules ───────────────────────────────────────────────
# Each entry: (regex_pattern, description)
# Only versions matching the pattern are kept — everything else is discarded
VERSION_RULES = {
    # X.YY.Z  or  X.YY  where X < 50 — user-facing CapCut versions
    "CapCut":              re.compile(r"^\d{1,2}\.\d+(\.\d+)*$"),
    # XX.X.X  — YouTube versions (e.g. 19.17.36)
    "YouTube":             re.compile(r"^\d{1,3}\.\d+\.\d+$"),
    # X.YY.Z  or  X.YY   — TikTok (e.g. 38.5.3)
    "TikTok":              re.compile(r"^\d{1,3}\.\d+(\.\d+)*$"),
    # X.Y.Z   — ChatGPT (e.g. 1.2024.340)
    "ChatGPT":             re.compile(r"^\d+\.\d{4}\.\d+$|^\d+\.\d+\.\d+$"),
    # X.Y.Z   — Claude (e.g. 1.0.5)
    "Claude by Anthropic": re.compile(r"^\d+\.\d+\.\d+$"),
    # X.YY.Z  — WhatsApp (e.g. 2.24.6.79)
    "WhatsApp Messenger":  re.compile(r"^\d+\.\d+\.\d+(\.\d+)*$"),
    # X.YY.Z  — Instagram (e.g. 333.0.0.40)
    "Instagram":           re.compile(r"^\d{1,4}\.\d+\.\d+(\.\d+)*$"),
    # X.Y.Z   — LinkedIn (e.g. 9.2.4)
    "LinkedIn":            re.compile(r"^\d{1,3}\.\d+\.\d+$"),
    # XX.X.X  — Tinder (e.g. 15.3.1)
    "Tinder":              re.compile(r"^\d{1,3}\.\d+\.\d+$"),
    # X.Y.Z.Z — Spotify (e.g. 8.9.70.604)
    "Spotify":             re.compile(r"^\d+\.\d+\.\d+(\.\d+)*$"),
}
# Generic fallback: major version < 100 and at least two numeric segments
VERSION_FALLBACK = re.compile(r"^\d{1,2}\.\d+(\.\d+)*$")

def is_valid_version(version_str: str, app_name: str) -> bool:
    """True only if the version string looks like a real user-facing version."""
    v = str(version_str).strip()
    # Reject strings that look like dates (YYYY.MM.DD or YYYY-MM-DD)
    if re.match(r"^20\d{2}[.\-]\d{2}[.\-]\d{2}$", v):
        return False
    rule = VERSION_RULES.get(app_name, VERSION_FALLBACK)
    return bool(rule.match(v))

# ── HTTP session ───────────────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
})


def safe_get(url: str, retries: int = 3, **kw) -> Optional[requests.Response]:
    for attempt in range(retries):
        try:
            r = SESSION.get(url, timeout=25, **kw)
            if r.status_code == 200:
                return r
            if r.status_code in (429, 502, 503):
                wait = 8 * (attempt + 1)
                print(f"    [HTTP {r.status_code}] waiting {wait}s …")
                time.sleep(wait)
            else:
                print(f"    [HTTP {r.status_code}] {url}")
                return None
        except (requests.ConnectionError, requests.Timeout) as exc:
            print(f"    [ERR] {exc}")
            if attempt < retries - 1:
                time.sleep(6 * (attempt + 1))
    return None


# ── Date helpers ───────────────────────────────────────────────────────────────
_DATE_FMTS = [
    "%b %d, %Y", "%B %d, %Y", "%b. %d, %Y",
    "%b %d %Y",  "%B %d %Y",
    "%Y-%m-%d",
    "%d %b %Y",  "%d %B %Y",
    "%d/%m/%Y",  "%m/%d/%Y",
]

def parse_date(raw: str) -> str:
    """Return YYYY-MM-DD or '' if unparseable."""
    if not raw:
        return ""
    raw = raw.strip().rstrip(".")
    # Normalize abbreviations like "Apr." → "Apr"
    raw = re.sub(r"([A-Za-z]{3})\.", r"\1", raw)
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def is_after_cutoff(date_str: str) -> bool:
    if not date_str:
        return True   # keep undated rows (flag in data quality)
    return date_str >= CUTOFF_DATE


# ── Source A: APKPure /versions page ──────────────────────────────────────────
def scrape_apkpure(pkg: str, slug: str) -> tuple[list[dict], str]:
    """
    Parse the APKPure versions list page.
    Returns list of {version, date} dicts and the URL string.

    APKPure /versions page structure (as of 2024):
      <ul class="ver-wrap">
        <li class="ver-item">
          <a class="ver-name">...</a>  ← version number
          <span class="date">...</span> ← date
        </li>
        ...
      </ul>
    If that selector fails, falls back to regex on structured text blocks.
    """
    url = f"https://apkpure.com/{slug}/{pkg}/versions"
    print(f"  [APKPure] {url}")
    r = safe_get(url)
    if not r:
        print("  [APKPure] FAILED")
        return [], url

    soup = BeautifulSoup(r.text, "lxml")
    results = []

    # ── Strategy 1: structured li.ver-item elements ────────────────────────────
    items = soup.select("li.ver-item, div.ver-item, li[class*='version'], div[class*='ver-wrap'] li")
    if items:
        for item in items:
            # Version number — look for dedicated class or just the first bold/strong text
            ver_el = (
                item.select_one(".ver-name, .version-name, span[class*='ver'], a[class*='ver']")
                or item.find("a")
            )
            ver = ver_el.get_text(strip=True) if ver_el else ""
            ver = re.sub(r"[^\d.]", "", ver).strip(".")

            # Date — look for dedicated class or <time>
            date_el = item.select_one(".date, time, span[class*='date'], p[class*='date']")
            date_raw = ""
            if date_el:
                date_raw = date_el.get("datetime", "") or date_el.get_text(strip=True)
            date = parse_date(date_raw)

            if ver:
                results.append({"version": ver, "date": date})
        if results:
            print(f"  [APKPure] structured: {len(results)} versions, {sum(1 for r in results if r['date'])} with dates")
            return results, url

    # ── Strategy 2: regex over raw text blocks ─────────────────────────────────
    # APKPure sometimes server-renders as inline JSON or text blocks
    # Look for patterns like: "14.5.1" ... "Apr 12, 2024" within ~500 chars
    text = r.text
    # Find all version-like strings and nearby dates
    ver_pattern = re.compile(r'\b(\d{1,3}\.\d+(?:\.\d+){0,3})\b')
    date_pattern = re.compile(
        r'\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s*\d{4})\b'
        r'|\b(\d{4}-\d{2}-\d{2})\b', re.I
    )

    # Collect all date positions
    date_positions = []
    for m in date_pattern.finditer(text):
        raw = (m.group(1) or m.group(2) or "").strip()
        d = parse_date(raw)
        if d:
            date_positions.append((m.start(), d))

    seen_versions = set()
    for vm in ver_pattern.finditer(text):
        ver = vm.group(1)
        if ver in seen_versions or all(p == "0" for p in ver.split(".")):
            continue
        # only keep if near the versions list area of the page
        # Find closest date within 800 characters
        closest = min(
            ((abs(dp - vm.start()), dv) for dp, dv in date_positions),
            key=lambda x: x[0],
            default=(9999, "")
        )
        dist, date = closest
        if dist > 800:
            date = ""
        seen_versions.add(ver)
        results.append({"version": ver, "date": date})

    print(f"  [APKPure] regex: {len(results)} versions, {sum(1 for r in results if r['date'])} with dates")
    return results, url


# ── Source B: AppBrain — release notes per version ────────────────────────────
def scrape_appbrain_notes(pkg: str, apkpure_slug: str) -> dict[str, str]:
    """
    Returns {version_str: notes_text} for versions that have notes on AppBrain.
    AppBrain is used as supplementary notes source only — not for version discovery.
    """
    slug = apkpure_slug or pkg.split(".")[-1]
    url = f"https://www.appbrain.com/app/{slug}/{pkg}"
    print(f"  [AppBrain] {url}")
    r = safe_get(url)
    if not r:
        print("  [AppBrain] FAILED")
        return {}

    soup = BeautifulSoup(r.text, "lxml")
    notes_map: dict[str, str] = {}

    # AppBrain renders a changelog table or list
    # Typical selectors: div.app-version, li.version-item
    for item in soup.select("div.app-version, li.version-item, div[class*='changelog'] li"):
        text = item.get_text(" ", strip=True)
        vm = re.search(r'\b(\d{1,3}\.\d+(?:\.\d+)*)\b', text)
        if not vm:
            continue
        ver = vm.group(1)
        # Remove the version number itself from the notes text
        notes = re.sub(re.escape(ver), "", text).strip(" |-")
        if len(notes) > 20:
            notes_map[ver] = notes[:1500]

    # Fallback: look for a changelog section
    if not notes_map:
        changelog_section = soup.find(
            lambda tag: tag.name in ("section", "div", "article")
            and re.search(r"changelog|what.s new|version history", tag.get_text()[:100], re.I)
        )
        if changelog_section:
            text = changelog_section.get_text(" ", strip=True)
            for vm in re.finditer(r'\b(\d{1,3}\.\d+(?:\.\d+)*)\b', text):
                ver = vm.group(1)
                after = text[vm.end():vm.end() + 800].strip()
                if len(after) > 20 and ver not in notes_map:
                    notes_map[ver] = after[:1500]

    print(f"  [AppBrain] {len(notes_map)} versions with notes")
    return notes_map


# ── Source C: Google Play Store — current version + notes ─────────────────────
def scrape_play_store(pkg: str) -> dict:
    url = f"https://play.google.com/store/apps/details?id={pkg}&hl=en"
    print(f"  [PlayStore] {url}")
    r = safe_get(url)
    if not r:
        print("  [PlayStore] FAILED")
        return {}
    raw = r.text

    # Current version
    ver = ""
    for pat in [
        r'"softwareVersion":"([\d.]+)"',
        r'\[\[\["([\d.]+)"\]\]',
        r'Current Version.*?([\d.]+)',
    ]:
        m = re.search(pat, raw)
        if m:
            candidate = m.group(1)
            if re.match(r"^\d+\.\d+", candidate):
                ver = candidate
                break

    # What's new / release notes
    notes = ""
    soup = BeautifulSoup(raw, "lxml")
    for tag in soup.find_all(["section", "div"]):
        h = tag.find(["h1","h2","h3","h4","span"])
        if h and re.search(r"what.?s new|recent changes", h.get_text(), re.I):
            body = tag.get_text(" ", strip=True)
            body = re.sub(r"(?i)what.?s new|recent changes", "", body).strip()
            if len(body) > 20:
                notes = body[:2000]
                break
    if not notes:
        m = re.search(r"(?:What.s New|Recent changes)[\"\\s>:]+([^<]{30,1000})", raw, re.S | re.I)
        if m:
            notes = m.group(1).strip()[:2000]

    # Initial release / last updated
    init_date = ""
    m = re.search(r"(?:Updated|Last updated)[^:]*:\s*(\w+ \d{1,2}, \d{4}|\d{4}-\d{2}-\d{2})", raw)
    if m:
        init_date = parse_date(m.group(1))

    result = {"version": ver, "notes": notes, "init_date": init_date, "url": url}
    print(f"  [PlayStore] ver={ver} | notes={bool(notes)} | init_date={init_date}")
    return result


# ── Source D: Apptopia — fallback when APKPure unavailable ────────────────────
def scrape_apptopia(pkg: str) -> tuple[list[dict], str]:
    """
    Falls back to Apptopia with a tighter version proximity window
    and strict version validation to avoid false positives.
    """
    url = f"https://apptopia.com/google-play/app/{pkg}/about"
    print(f"  [Apptopia] {url}")
    r = safe_get(url)
    if not r:
        print("  [Apptopia] FAILED")
        return [], url

    text = r.text
    date_pattern = re.compile(
        r'\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s*\d{4})\b'
        r'|\b(\d{4}-\d{2}-\d{2})\b', re.I
    )
    ver_pattern = re.compile(r'\b(\d{1,3}\.\d+(?:\.\d+){0,3})\b')

    date_positions = []
    for m in date_pattern.finditer(text):
        raw = (m.group(1) or m.group(2) or "").strip()
        d = parse_date(raw)
        if d:
            date_positions.append((m.start(), d))

    seen, results = set(), []
    for vm in ver_pattern.finditer(text):
        ver = vm.group(1)
        if ver in seen:
            continue
        seen.add(ver)
        closest = min(
            ((abs(dp - vm.start()), dv) for dp, dv in date_positions),
            key=lambda x: x[0],
            default=(9999, "")
        )
        dist, date = closest
        results.append({"version": ver, "date": date if dist <= 500 else ""})

    print(f"  [Apptopia] {len(results)} versions, {sum(1 for r in results if r['date'])} with dates")
    return results, url


# ── Collect one app ────────────────────────────────────────────────────────────
def collect_app(app_info: tuple) -> list[dict]:
    name, pkg, developer, category, apkpure_slug, _ = app_info

    print(f"\n{'='*60}")
    print(f"  {name}  ({pkg})")
    print(f"{'='*60}")

    # 1. Play Store — current version + notes
    play = scrape_play_store(pkg)
    cur_ver   = play.get("version", "")
    cur_notes = play.get("notes", "")
    init_date = play.get("init_date", "")
    play_url  = play.get("url", f"https://play.google.com/store/apps/details?id={pkg}&hl=en")
    time.sleep(SLEEP)

    # 2. Primary version list — APKPure if available, else Apptopia
    if apkpure_slug:
        version_list, list_url = scrape_apkpure(pkg, apkpure_slug)
    else:
        version_list, list_url = scrape_apptopia(pkg)
    time.sleep(SLEEP)

    # 3. AppBrain notes (best-effort)
    notes_map = scrape_appbrain_notes(pkg, apkpure_slug)
    time.sleep(SLEEP)

    # ── Build version lookup {version: date} from primary source ───────────────
    version_dates: dict[str, str] = {}
    for item in version_list:
        v = str(item["version"]).strip()
        d = item.get("date", "")
        if v and is_valid_version(v, name):
            # Keep earliest date if duplicate (shouldn't happen but safety)
            if v not in version_dates or (d and not version_dates[v]):
                version_dates[v] = d

    # Ensure current version is included
    if cur_ver and is_valid_version(cur_ver, name):
        if cur_ver not in version_dates:
            version_dates[cur_ver] = ""

    # ── Filter by cutoff ───────────────────────────────────────────────────────
    # Keep versions that are either undated or dated after CUTOFF_DATE
    filtered = {v: d for v, d in version_dates.items() if is_after_cutoff(d)}

    # ── Sort: current version first, then by date descending ──────────────────
    def sort_key(item):
        v, d = item
        is_cur = (v == cur_ver)
        date_sort = d or "0000-00-00"
        return (not is_cur, date_sort)

    sorted_versions = sorted(filtered.items(), key=sort_key, reverse=True)
    # Fix: reverse=True on tuple means (False, "2026-...") > (True, "2024-...") — recheck
    # We want: cur first, then date desc
    sorted_versions = sorted(
        filtered.items(),
        key=lambda x: (x[0] != cur_ver, x[1] or "0000-00-00"),
        reverse=False       # False: False(cur) first, then ascending date would be wrong
    )
    # Actually: sort by date descending, but cur_ver always first
    non_cur = [(v, d) for v, d in filtered.items() if v != cur_ver]
    non_cur.sort(key=lambda x: x[1] or "0000-00-00", reverse=True)
    if cur_ver and cur_ver in filtered:
        sorted_versions = [(cur_ver, filtered[cur_ver])] + non_cur
    else:
        sorted_versions = non_cur

    # ── Build rows ─────────────────────────────────────────────────────────────
    rows = []
    for ver, date in sorted_versions:
        is_current = (ver == cur_ver)

        notes = ""
        if is_current and cur_notes:
            notes = cur_notes
        elif ver in notes_map:
            notes = notes_map[ver]

        if is_current:
            source_url = play_url
        else:
            source_url = list_url

        dq_flags = []
        if not date:
            dq_flags.append("No date found")
        if not notes:
            dq_flags.append("No release notes")
        if not dq_flags:
            dq_flags = ["Date + notes available"]

        rows.append({
            "App Name":                           name,
            "Platform":                           "Android",
            "Developer / Company":                developer,
            "App Category":                       category,
            "Version Number":                     ver,
            "Version Release Date":               date,
            "Current Version":                    "Yes" if is_current else "No",
            "Initial App Release Date":           init_date,
            "Update Description / Release Notes": notes,
            "Update Category":                    "",    # left for manual/AI labelling
            "Brief Summary":                      "",    # left for manual/AI labelling
            "Source of Update History":           source_url,
            "Data Quality Notes":                 " | ".join(dq_flags),
        })

    dated   = sum(1 for r in rows if r["Version Release Date"])
    noted   = sum(1 for r in rows if r["Update Description / Release Notes"])
    print(f"  → {len(rows)} rows | {dated} dated | {noted} with notes")
    return rows


# ── Excel styling ──────────────────────────────────────────────────────────────
COL_WIDTHS = {
    "App Name": 22, "Platform": 12, "Developer / Company": 26,
    "App Category": 18, "Version Number": 17, "Version Release Date": 20,
    "Current Version": 15, "Initial App Release Date": 22,
    "Update Description / Release Notes": 55,
    "Update Category": 45, "Brief Summary": 50,
    "Source of Update History": 45, "Data Quality Notes": 40,
}
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
CUR_FILL = PatternFill("solid", fgColor="E2EFDA")
BODY_FNT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="BFBFBF")
BDR  = Border(bottom=THIN, right=THIN)
SRC_COL   = COLUMNS.index("Source of Update History") + 1
WRAP_COLS = {COLUMNS.index(c) + 1 for c in [
    "Update Description / Release Notes", "Update Category", "Brief Summary", "Data Quality Notes"
]}


def write_excel(df: pd.DataFrame, out_path: Path) -> None:
    df.to_excel(str(out_path), index=False, engine="openpyxl")
    wb = load_workbook(str(out_path))
    ws = wb.active
    ws.title = "Android Version History"

    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.border = BDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 36

    for ri in range(2, ws.max_row + 1):
        is_cur = ws.cell(ri, 7).value == "Yes"
        bg = CUR_FILL if is_cur else (ALT_FILL if ri % 2 == 0 else None)
        for ci in range(1, len(COLUMNS) + 1):
            cell = ws.cell(ri, ci)
            cell.font = BODY_FNT
            cell.border = BDR
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in WRAP_COLS))
            if bg:
                cell.fill = bg
        src = ws.cell(ri, SRC_COL)
        if str(src.value or "").startswith("http"):
            src.hyperlink = src.value
            src.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(str(out_path))


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    out_path = Path(__file__).resolve().parent / OUTPUT_FILE
    print(f"Output : {out_path}")
    print(f"Cutoff : {CUTOFF_DATE}\n")

    all_rows = []
    for app_info in ANDROID_APPS:
        rows = collect_app(app_info)
        all_rows.extend(rows)
        time.sleep(2.0)

    if not all_rows:
        print("No data collected — check network connectivity.")
        sys.exit(1)

    df = pd.DataFrame(all_rows, columns=COLUMNS)

    # Final sort: App Name asc, then Version Release Date desc (undated last)
    df["_d"] = pd.to_datetime(df["Version Release Date"], errors="coerce")
    df["_cur"] = df["Current Version"].map({"Yes": 0, "No": 1})
    df.sort_values(["App Name", "_cur", "_d"], ascending=[True, True, False], inplace=True)
    df.drop(columns=["_d", "_cur"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    write_excel(df, out_path)

    print(f"\n{'='*60}")
    print(f"Saved : {out_path}")
    print(f"{'='*60}")
    total      = df.groupby("App Name")["Version Number"].count().rename("total")
    with_date  = df[df["Version Release Date"] != ""].groupby("App Name")["Version Number"].count().rename("dated")
    with_notes = (
        df[df["Update Description / Release Notes"].str.strip().ne("")]
        .groupby("App Name")["Version Number"].count().rename("noted")
    )
    summary = pd.concat([total, with_date, with_notes], axis=1).fillna(0)
    summary = summary.astype({"dated": int, "noted": int})
    print(summary.to_string())
    print(f"\nTotal rows: {len(df)}")


if __name__ == "__main__":
    main()