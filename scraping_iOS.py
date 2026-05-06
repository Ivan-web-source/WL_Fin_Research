import json, re, sys, time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INPUT_FILE   = "ios_app_version_history.xlsx"
OUTPUT_FILE  = "ios_app_version_history_v3.xlsx"
CUTOFF_DATE  = "2023-05-01"
PROX_MAX_DIST = 600
SLEEP        = 1.5

IOS_APPS = [
    ("YouTube",             "544007664",  "Google",               "Photo & Video",     "iOS"),
    ("TikTok",              "835599320",  "TikTok Ltd.",           "Entertainment",     "iOS"),
    ("ChatGPT",             "6448311069", "OpenAI OpCo, LLC",      "Productivity",      "iOS"),
    ("Claude by Anthropic", "6473753684", "Anthropic PBC",         "Productivity",      "iOS"),
    ("WhatsApp Messenger",  "310633997",  "WhatsApp Inc.",         "Social Networking", "iOS"),
    ("CapCut",              "1500855883", "Bytedance Pte. Ltd",    "Photo & Video",     "iOS"),
    ("Instagram",           "389801252",  "Instagram, Inc.",       "Photo & Video",     "iOS"),
    ("LinkedIn",            "288429040",  "LinkedIn Corporation",  "Business",          "iOS"),
    ("Tinder",              "547702041",  "Tinder LLC",            "Lifestyle",         "iOS"),
    ("Spotify",             "324684580",  "Spotify",               "Music",             "iOS"),
]

COLUMNS = [
    "App Name", "Platform", "Developer / Company", "App Category",
    "Version Number", "Version Release Date", "Current Version",
    "Initial App Release Date", "Update Description / Release Notes",
    "Update Category", "Brief Summary",
    "Source of Update History", "Data Quality Notes",
]

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
})

def safe_get(url, retries=3, **kw):
    for attempt in range(retries):
        try:
            r = SESSION.get(url, timeout=25, **kw)
            if r.status_code == 200:
                return r
            if r.status_code in (429, 502, 503):
                time.sleep(6 * (attempt + 1))
        except (requests.ConnectionError, requests.Timeout):
            if attempt < retries - 1:
                time.sleep(6 * (attempt + 1))
    return None

_DATE_PATTERNS = [
    (re.compile(r'\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4})\b', re.I),
     ["%b %d, %Y", "%B %d, %Y", "%b. %d, %Y", "%b %d %Y", "%B %d %Y"]),
    (re.compile(r'\b(\d{4}-\d{2}-\d{2})\b'),
     ["%Y-%m-%d"]),
    (re.compile(r'\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4})\b', re.I),
     ["%d %b %Y", "%d %B %Y"]),
]

def parse_date_str(s):
    for _, fmts in _DATE_PATTERNS:
        for fmt in fmts:
            try:
                return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None

def find_dates_pos(text):
    results = []
    seen = set()
    for rx, fmts in _DATE_PATTERNS:
        for m in rx.finditer(text):
            if m.start() in seen:
                continue
            for fmt in fmts:
                try:
                    d = datetime.strptime(m.group(1).strip(), fmt).strftime("%Y-%m-%d")
                    results.append((m.start(), d))
                    seen.add(m.start())
                    break
                except ValueError:
                    continue
    return sorted(results, key=lambda x: x[0])

def find_versions_pos(text):
    results, seen = [], set()
    for m in re.finditer(r'\b(\d+\.\d+(?:\.\d+)*)\b', text):
        v = m.group(1)
        if all(p == "0" for p in v.split(".")) or v in seen:
            continue
        seen.add(v)
        results.append((m.start(), v))
    return results

def proximity_pair(text, max_dist=PROX_MAX_DIST):
    versions = find_versions_pos(text)
    dates    = find_dates_pos(text)
    if not dates:
        return [(v, "") for _, v in versions]
    pairs = []
    for vpos, ver in versions:
        nearest_date, nearest_dist = "", float("inf")
        for dpos, dval in dates:
            dist = abs(dpos - vpos)
            if dist < nearest_dist:
                nearest_dist, nearest_date = dist, dval
        pairs.append((ver, nearest_date if nearest_dist <= max_dist else ""))
    return pairs

def parse_iso(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip()[:25], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def is_after_cutoff(date_str):
    if not date_str:
        return True  # undated rows kept
    try:
        return date_str >= CUTOFF_DATE
    except Exception:
        return True

def source_apptopia(app_id):
    url = f"https://apptopia.com/ios/app/{app_id}/about"
    print(f"  [Apptopia] {url}")
    r = safe_get(url)
    if not r:
        return [], url
    pairs = proximity_pair(r.text)
    dated = sum(1 for _, d in pairs if d)
    print(f"  [Apptopia] {len(pairs)} versions, {dated} with dates")
    return pairs, url

def source_itunes(app_id):
    r = safe_get(f"https://itunes.apple.com/lookup?id={app_id}&country=us")
    if not r:
        return {}
    try:
        data = r.json().get("results", [{}])[0]
        return {
            "version":   data.get("version", ""),
            "date":      parse_iso(data.get("currentVersionReleaseDate", "")),
            "notes":     data.get("releaseNotes", ""),
            "init_date": parse_iso(data.get("releaseDate", "")),
        }
    except Exception:
        return {}

def collect_app(app_info, existing_df):
    name, app_id, developer, category, platform = app_info

    print(f"\n{'='*60}")
    print(f"  {name}  (id={app_id})")
    print(f"{'='*60}")

    itunes = source_itunes(app_id)
    cur_ver   = itunes.get("version", "")
    cur_date  = itunes.get("date", "")
    cur_notes = itunes.get("notes", "")
    init_date = itunes.get("init_date", "")
    itunes_url = f"https://itunes.apple.com/lookup?id={app_id}&country=us"
    time.sleep(SLEEP)

    apptopia_pairs, apptopia_url = source_apptopia(app_id)
    time.sleep(SLEEP)

    apptopia_dates = {v: d for v, d in apptopia_pairs if v and d}

    existing_rows = existing_df[existing_df["App Name"] == name]
    existing_dates = {}
    existing_notes = {}
    for _, row in existing_rows.iterrows():
        v = str(row.get("Version Number", ""))
        d = str(row.get("Version Release Date", ""))
        n = str(row.get("Update Description / Release Notes", ""))
        if v and v != "nan":
            if d and d not in ("nan", "NaT", ""):
                existing_dates[v] = d
            if n and n != "nan":
                existing_notes[v] = n

    all_versions, seen = [], set()
    if cur_ver:
        all_versions.append(cur_ver)
        seen.add(cur_ver)
    for v in existing_rows["Version Number"].astype(str):
        if v and v not in seen and v != "nan":
            all_versions.append(v)
            seen.add(v)
    for v, _ in apptopia_pairs:
        if v and v not in seen:
            all_versions.append(v)
            seen.add(v)

    rows = []
    for ver in all_versions:
        is_current = ver == cur_ver

        if is_current and cur_date:
            date, date_src = cur_date, "iTunes API"
        elif ver in existing_dates:
            date, date_src = existing_dates[ver], "existing"
        elif ver in apptopia_dates:
            date, date_src = apptopia_dates[ver], "Apptopia"
        else:
            date, date_src = "", ""

        if not is_after_cutoff(date) and date:
            continue

        if is_current and cur_notes:
            notes = cur_notes
        elif ver in existing_notes:
            notes = existing_notes[ver]
        else:
            notes = ""

        source_url = itunes_url if is_current else apptopia_url

        dq_flags = []
        if not date:
            dq_flags.append("No date found")
        else:
            dq_flags.append(f"Date: {date_src}")
        if not notes:
            dq_flags.append("No release notes")

        rows.append({
            "App Name":                           name,
            "Platform":                           platform,
            "Developer / Company":                developer,
            "App Category":                       category,
            "Version Number":                     ver,
            "Version Release Date":               date,
            "Current Version":                    "Yes" if is_current else "No",
            "Initial App Release Date":           init_date,
            "Update Description / Release Notes": notes,
            "Update Category":                    "",
            "Brief Summary":                      "",
            "Source of Update History":           source_url,
            "Data Quality Notes":                 " | ".join(dq_flags),
        })

    print(f"  -> {len(rows)} rows | {sum(1 for r in rows if r['Version Release Date'])} with dates | {sum(1 for r in rows if r['Update Description / Release Notes'])} with notes")
    return rows

COL_WIDTHS = {
    "App Name": 22, "Platform": 10, "Developer / Company": 26,
    "App Category": 18, "Version Number": 17, "Version Release Date": 20,
    "Current Version": 15, "Initial App Release Date": 22,
    "Update Description / Release Notes": 55,
    "Update Category": 45, "Brief Summary": 50,
    "Source of Update History": 45, "Data Quality Notes": 40,
}
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
CUR_FILL = PatternFill("solid", fgColor="E2EFDA")
BODY_FNT = Font(name="Calibri", size=10)
THIN     = Side(style="thin", color="BFBFBF")
BDR      = Border(bottom=THIN, right=THIN)
SRC_COL  = COLUMNS.index("Source of Update History") + 1

def write_excel(df, out_path):
    df.to_excel(str(out_path), index=False, engine="openpyxl")
    wb = load_workbook(str(out_path))
    ws = wb.active
    ws.title = "iOS Version History"
    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 36
    wrap_cols = {COLUMNS.index(c)+1 for c in ["Update Description / Release Notes", "Update Category", "Brief Summary", "Data Quality Notes"]}
    for ri in range(2, ws.max_row + 1):
        is_cur = ws.cell(ri, 7).value == "Yes"
        bg = CUR_FILL if is_cur else (ALT_FILL if ri % 2 == 0 else None)
        for ci in range(1, len(COLUMNS) + 1):
            cell = ws.cell(ri, ci)
            cell.font = BODY_FNT; cell.border = BDR
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in wrap_cols))
            if bg:
                cell.fill = bg
        src = ws.cell(ri, SRC_COL)
        if str(src.value or "").startswith("http"):
            src.hyperlink = src.value
            src.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(str(out_path))

def main():
    script_dir = Path(__file__).resolve().parent
    input_path = script_dir / INPUT_FILE
    out_path   = script_dir / OUTPUT_FILE

    if not input_path.exists():
        print(f"ERROR: {input_path} not found.")
        sys.exit(1)

    existing_df = pd.read_excel(str(input_path))
    print(f"Loaded {len(existing_df)} existing rows | Cutoff: {CUTOFF_DATE}\n")

    all_rows = []
    for app_info in IOS_APPS:
        rows = collect_app(app_info, existing_df)
        all_rows.extend(rows)
        time.sleep(2.0)

    df = pd.DataFrame(all_rows, columns=COLUMNS)
    df["_d"] = pd.to_datetime(df["Version Release Date"], errors="coerce")
    df.sort_values(["App Name", "_d"], ascending=[True, False], inplace=True)
    df.drop(columns=["_d"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    write_excel(df, out_path)

    print(f"\n{'='*60}")
    print(f"Saved: {out_path}")
    print(f"{'='*60}")
    total      = df.groupby("App Name")["Version Number"].count().rename("total")
    with_date  = df[df["Version Release Date"] != ""].groupby("App Name")["Version Number"].count().rename("dated")
    with_notes = df[df["Update Description / Release Notes"] != ""].groupby("App Name")["Version Number"].count().rename("noted")
    print(pd.concat([total, with_date, with_notes], axis=1).fillna(0).astype({"dated": int, "noted": int}).to_string())
    print(f"\nTotal rows: {len(df)}")

if __name__ == "__main__":
    main()